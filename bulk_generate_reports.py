from __future__ import annotations

import json
import subprocess
from datetime import date, datetime, timedelta, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ================= CONFIG =================
DSN = "pos"
UID = "db"
PWD = "db"

BUSINESS_START = "06:30"

REPO_ROOT = Path(__file__).resolve().parent
REPORTS_DIR = REPO_ROOT / "reports"
INDEX_JSON = REPO_ROOT / "report_index.json"

TODAY = date.today()
START_DATE = date(TODAY.year - 1, 12, 1)
END_DATE = TODAY

DO_GIT_PUSH = True
GIT_MESSAGE = "Update reports with HTML and XLSX"
# ==========================================


def d0(v) -> Decimal:
    try:
        if v is None:
            return Decimal("0")
        if isinstance(v, Decimal):
            return v
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def fmt2(v) -> str:
    return f"{d0(v):,.2f}"


def fmt_int(v) -> str:
    try:
        return f"{int(d0(v)):,}"
    except Exception:
        return "0"


def business_window(day: date):
    hh, mm = map(int, BUSINESS_START.split(":"))
    start = datetime.combine(day, time(hh, mm))
    end = start + timedelta(days=1)
    return start, end


def connect_db():
    return pyodbc.connect(f"DSN={DSN};UID={UID};PWD={PWD};", autocommit=True)


def ensure_dir(path: Path):
    path.mkdir(parents=True, exist_ok=True)


def write_text(path: Path, content: str):
    ensure_dir(path.parent)
    path.write_text(content, encoding="utf-8")


def bright_shell(title: str, subtitle: str, table_html: str) -> str:
    return f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>{title}</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  body {{
    margin:0;
    padding:12px;
    font-family: Arial, sans-serif;
    background:#c9edf7;
    color:#000;
    font-size:13px;
  }}
  .header {{
    background:#d9f4fb;
    border:1px solid #999;
    padding:8px 10px;
    margin-bottom:8px;
  }}
  h2 {{
    margin:0 0 4px 0;
    font-size:18px;
    color:#00334d;
  }}
  .meta {{
    font-size:12px;
    color:#333;
  }}
  table {{
    border-collapse:collapse;
    width:100%;
    background:#c9edf7;
    table-layout:auto;
  }}
  th {{
    background:#d9f4fb;
    border:1px solid #999;
    padding:5px 8px;
    text-align:center;
    font-weight:bold;
    color:#000;
  }}
  td {{
    border:1px solid #999;
    padding:4px 8px;
    color:#000;
  }}
  .right {{
    text-align:right;
    font-family: Arial, sans-serif;
  }}
  .left {{
    text-align:left;
  }}
  .total {{
    color:red;
    font-weight:bold;
  }}
  .section {{
    font-weight:bold;
  }}
</style>
</head>
<body>
  <div class="header">
    <h2>{title}</h2>
    <div class="meta">{subtitle}</div>
  </div>
  {table_html}
</body>
</html>
"""


def write_xlsx(path: Path, sheet_name: str, headers: list[str], rows: list[list], total_labels: set[str]):
    ensure_dir(path.parent)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="D9F4FB")
    body_fill = PatternFill("solid", fgColor="C9EDF7")
    total_font = Font(bold=True, color="FF0000")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(r)

    for row in ws.iter_rows(min_row=2):
        label = str(row[0].value or "")
        is_total = label in total_labels or label.upper() == "TOTAL"

        for idx, cell in enumerate(row, start=1):
            cell.fill = body_fill
            cell.border = border
            if is_total:
                cell.font = total_font
            if idx > 1:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 3, 45)

    wb.save(path)


# ================= DAILY SUMMARY =================
def build_daily_summary(conn, start_dt, end_dt, report_date: date):
    cur = conn.cursor()

    sql = """
    SELECT
      TransType,
      GroupTransType,
      ReceiptN,
      SubCategoryID,
      Amount,
      DiscountAmount,
      TaxInclude,
      Tax1Amount,
      Tax2Amount,
      Tax3Amount,
      Tax4Amount
    FROM Journal
    WHERE DateR >= ? AND DateR < ?
      AND Status = 0
      AND (TransType IN (101,102,103,104,311,501,780) OR GroupTransType IN (1,2))
    """

    rows = cur.execute(sql, start_dt, end_dt).fetchall()
    sales = [r for r in rows if int(r.TransType or 0) == 101]

    def tax_total(r):
        return d0(r.Tax1Amount) + d0(r.Tax2Amount) + d0(r.Tax3Amount) + d0(r.Tax4Amount)

    def tax_add_multiplier(r):
        return Decimal("1") - d0(r.TaxInclude)

    gross = sum(d0(r.Amount) + tax_total(r) * tax_add_multiplier(r) for r in sales)

    gst = sum(d0(r.Tax1Amount) * tax_add_multiplier(r) for r in sales)
    pst = sum(d0(r.Tax2Amount) * tax_add_multiplier(r) for r in sales)
    liq = sum(d0(r.Tax3Amount) * tax_add_multiplier(r) for r in sales)
    tax4 = sum(d0(r.Tax4Amount) * tax_add_multiplier(r) for r in sales)

    total_taxes = gst + pst + liq + tax4
    net = gross - total_taxes
    discount = sum(d0(r.DiscountAmount) for r in sales)

    customers = len({
        r.ReceiptN for r in rows
        if int(r.GroupTransType or 0) == 1 and r.ReceiptN is not None
    })

    avg_sale = net / Decimal(customers) if customers else Decimal("0")

    by_subcat = {}
    for r in sales:
        key = str(r.SubCategoryID or "").strip() or "UNSPECIFIED"
        net_line = d0(r.Amount) - tax_total(r) * d0(r.TaxInclude)
        by_subcat[key] = by_subcat.get(key, Decimal("0")) + net_line

    day_col = report_date.strftime("%B %d %a")
    total_labels = {
        "Total Sales:",
        "Net Total Sales",
        "Total taxes",
        "Total Sales",
    }

    html_rows = []
    xlsx_rows = []

    def add_row(label, value, is_total=False):
        cls = "total" if is_total else ""
        html_rows.append(
            f"<tr><td class='left {cls}'>{label}</td>"
            f"<td class='right {cls}'>{fmt2(value)}</td>"
            f"<td class='right {cls}'>{fmt2(value)}</td></tr>"
        )
        xlsx_rows.append([label, float(d0(value)), float(d0(value))])

    add_row("Total Sales:", gross, True)

    for k in sorted(by_subcat):
        add_row(k, by_subcat[k], False)

    add_row("Net Total Sales", net, True)
    add_row("GST 5%", gst)
    add_row("PST 7%", pst)
    add_row("LIQ TAX 10%", liq)

    if tax4 != 0:
        add_row("Tax4", tax4)

    add_row("Total taxes", total_taxes, True)
    add_row("Total Sales", gross, True)
    add_row("Discount", discount)

    html_rows.append(
        f"<tr><td class='left'>Customer count</td>"
        f"<td class='right'>{customers}</td>"
        f"<td class='right'>{customers}</td></tr>"
    )
    xlsx_rows.append(["Customer count", customers, customers])

    add_row("Average Sale", avg_sale)

    table_html = f"""
<table>
  <thead>
    <tr>
      <th>Description</th>
      <th>{day_col}</th>
      <th>Total</th>
    </tr>
  </thead>
  <tbody>
    {''.join(html_rows)}
  </tbody>
</table>
"""

    subtitle = f"Business window: {start_dt:%Y-%m-%d %H:%M:%S} → {end_dt:%Y-%m-%d %H:%M:%S}"
    html = bright_shell("Daily Summary", subtitle, table_html)

    return html, xlsx_rows, ["Description", day_col, "Total"], total_labels


# ================= CATEGORY REPORT =================
def build_category_report(conn, start_dt, end_dt, report_date: date):
    cur = conn.cursor()

    sql = """
    SELECT
      C.SubCategoryID AS GroupName,
      SUM((J.Amount) - (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (J.TaxInclude)) AS AmountNet,
      SUM((J.Amount) + (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (1 - J.TaxInclude)) AS AmountTaxIncl,
      SUM(J.Quantity) AS CategoryCount,
      COUNT(DISTINCT J.ReceiptN) AS Customers
    FROM Journal J
      LEFT OUTER JOIN Category C ON C.CategoryID = J.CategoryID
    WHERE
      J.DateR >= ? AND J.DateR <= ?
      AND J.Status = 0
      AND (1 - C.SalesFlag) = 1
      AND J.TransType IN (101,102,112,111)
    GROUP BY C.SubCategoryID
    ORDER BY C.SubCategoryID
    """

    rows = cur.execute(sql, start_dt, end_dt).fetchall()

    total_sql = """
    SELECT
      SUM((J.Amount) - (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (J.TaxInclude)) AS AmountNet,
      SUM((J.Amount) + (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (1 - J.TaxInclude)) AS AmountTaxIncl,
      SUM(J.Quantity) AS CategoryCount,
      COUNT(DISTINCT J.ReceiptN) AS Customers
    FROM Journal J
      LEFT OUTER JOIN Category C ON C.CategoryID = J.CategoryID
    WHERE
      J.DateR >= ? AND J.DateR <= ?
      AND J.Status = 0
      AND (1 - C.SalesFlag) = 1
      AND J.TransType IN (101,102,112,111)
    """

    total = cur.execute(total_sql, start_dt, end_dt).fetchone()

    html_rows = []
    xlsx_rows = []

    for r in rows:
        name = str(r.GroupName or "UNSPECIFIED").strip()
        amount = d0(r.AmountNet)
        amount_tax = d0(r.AmountTaxIncl)
        qty = int(d0(r.CategoryCount))
        customers = int(d0(r.Customers))

        html_rows.append(
            f"<tr>"
            f"<td class='left'>{name}</td>"
            f"<td class='right'>{fmt2(amount)}</td>"
            f"<td class='right'>{fmt2(amount_tax)}</td>"
            f"<td class='right'>{qty}</td>"
            f"<td class='right'>{customers}</td>"
            f"</tr>"
        )

        xlsx_rows.append([name, float(amount), float(amount_tax), qty, customers])

    html_rows.append(
        f"<tr>"
        f"<td class='left total'>TOTAL</td>"
        f"<td class='right total'>{fmt2(total.AmountNet)}</td>"
        f"<td class='right total'>{fmt2(total.AmountTaxIncl)}</td>"
        f"<td class='right total'>{fmt_int(total.CategoryCount)}</td>"
        f"<td class='right total'>{fmt_int(total.Customers)}</td>"
        f"</tr>"
    )

    xlsx_rows.append([
        "TOTAL",
        float(d0(total.AmountNet)),
        float(d0(total.AmountTaxIncl)),
        int(d0(total.CategoryCount)),
        int(d0(total.Customers)),
    ])

    table_html = f"""
<table>
  <thead>
    <tr>
      <th>Group</th>
      <th>Amount</th>
      <th>Amount (Taxes Included)</th>
      <th>Category Count</th>
      <th>Customers</th>
    </tr>
  </thead>
  <tbody>
    {''.join(html_rows)}
  </tbody>
</table>
"""

    subtitle = f"Business window: {start_dt:%Y-%m-%d %H:%M:%S} → {end_dt:%Y-%m-%d %H:%M:%S}"
    html = bright_shell("Category Report", subtitle, table_html)

    headers = ["Group", "Amount", "Amount (Taxes Included)", "Category Count", "Customers"]
    return html, xlsx_rows, headers, {"TOTAL"}


# ================= INDEX =================
def update_report_index(dates):
    payload = {
        "latest": dates[0] if dates else None,
        "dates": dates,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    write_text(INDEX_JSON, json.dumps(payload, indent=2))


# ================= GIT =================
def run_cmd(cmd):
    return subprocess.run(
        cmd,
        cwd=str(REPO_ROOT),
        text=True,
        capture_output=True
    )


def git_sync():
    print("Running git add/commit/pull/push...")

    r = run_cmd(["git", "add", "."])
    if r.returncode != 0:
        print(r.stdout)
        print(r.stderr)
        return

    r = run_cmd(["git", "commit", "-m", "Update generated reports"])
    if r.returncode != 0:
        msg = (r.stdout + r.stderr).lower()
        if "nothing to commit" in msg:
            print("Nothing to commit.")
        else:
            print(r.stdout)
            print(r.stderr)
            return

    r = run_cmd(["git", "pull", "--rebase", "origin", "main"])
    if r.returncode != 0:
        print("Git pull failed:")
        print(r.stdout)
        print(r.stderr)
        return

    r = run_cmd(["git", "push", "origin", "main"])
    if r.returncode != 0:
        print("Git push failed:")
        print(r.stdout)
        print(r.stderr)
        return

    print("Git push successful.")


# ================= MAIN =================
def main():
    ensure_dir(REPORTS_DIR)

    print(f"Generating reports from {START_DATE} to {END_DATE}")
    print(f"Repo: {REPO_ROOT}")

    dates = []

    with connect_db() as conn:
        d = START_DATE

        while d <= END_DATE:
            ds = d.strftime("%Y-%m-%d")
            dates.append(ds)
            folder = REPORTS_DIR / ds
            ensure_dir(folder)

            start_dt, end_dt = business_window(d)

            try:
                daily_html, daily_xlsx, daily_headers, daily_totals = build_daily_summary(conn, start_dt, end_dt, d)
                category_html, category_xlsx, category_headers, category_totals = build_category_report(conn, start_dt, end_dt, d)

                write_text(folder / "summary_daily.html", daily_html)
                write_text(folder / "category_report.html", category_html)

                write_xlsx(folder / "summary_daily.xlsx", "Daily Summary", daily_headers, daily_xlsx, daily_totals)
                write_xlsx(folder / "category_report.xlsx", "Category Report", category_headers, category_xlsx, category_totals)

                print(f"OK {ds}")

            except Exception as e:
                print(f"FAIL {ds}: {e}")

            d += timedelta(days=1)

    dates.sort(reverse=True)
    update_report_index(dates)

    if DO_GIT_PUSH:
        git_sync()

    print("DONE")


if __name__ == "__main__":
    main()